#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Fill the P1 manual audit workbook according to 临时/人工审查/5审.md.

Offline only: no API calls, no answer generation, and no R3a code changes.
"""

from __future__ import annotations

import csv
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


TRUE_ROWS = {"P1-0077"}
AMBIGUOUS_ROWS = {"P1-0036"}

USEFUL_SAME = {
    "P1-0005",
    "P1-0007",
    "P1-0008",
    "P1-0016",
    "P1-0018",
    "P1-0022",
    "P1-0024",
    "P1-0030",
    "P1-0049",
    "P1-0053",
    "P1-0059",
    "P1-0063",
    "P1-0082",
    "P1-0083",
    "P1-0084",
    "P1-0087",
    "P1-0088",
    "P1-0097",
    "P1-0098",
    "P1-0099",
    "P1-0104",
    "P1-0105",
    "P1-0106",
    "P1-0108",
    "P1-0109",
    "P1-0111",
    "P1-0112",
    "P1-0118",
}

USEFUL_PARTIAL = {
    "P1-0012",
    "P1-0013",
    "P1-0023",
    "P1-0025",
    "P1-0026",
    "P1-0027",
    "P1-0028",
    "P1-0029",
    "P1-0051",
    "P1-0079",
    "P1-0081",
    "P1-0091",
    "P1-0102",
    "P1-0113",
    "P1-0114",
}

MANUAL_FIELDS = [
    "人工_是否同主题",
    "人工_问题前提是否抽取正确",
    "人工_候选记忆是否明确反驳问题前提",
    "人工_是否只是相关但不反驳",
    "人工_是否只是缺失信息而非反证",
    "人工_是否存在时间不匹配",
    "人工_是否足以支持答案",
    "人工_是否应该触发R3a",
    "人工_最终标签",
    "人工_错误来源",
    "人工_备注",
]

VALIDATION_CHOICES = {
    "人工_是否同主题": ["是", "否", "部分"],
    "人工_问题前提是否抽取正确": ["是", "否", "部分", "无法判断"],
    "人工_候选记忆是否明确反驳问题前提": ["是", "否", "模糊", "无法判断"],
    "人工_是否只是相关但不反驳": ["是", "否", "部分"],
    "人工_是否只是缺失信息而非反证": ["是", "否", "部分"],
    "人工_是否存在时间不匹配": ["是", "否", "不确定"],
    "人工_是否足以支持答案": ["是", "否", "部分", "无法判断"],
    "人工_是否应该触发R3a": ["是", "否", "不确定"],
    "人工_最终标签": ["真反证", "模糊反证", "非反证但有用", "非反证", "无法判断"],
    "人工_错误来源": [
        "R3a漏触发",
        "候选发现器误报",
        "问题前提抽取错误",
        "槽位归类错误",
        "时间判断问题",
        "明确反驳判断问题",
        "检索证据不足",
        "不属于R3a",
        "无法判断",
    ],
}


def judgement(row_id: str) -> dict[str, str]:
    if row_id in TRUE_ROWS:
        return {
            "人工_是否同主题": "是",
            "人工_问题前提是否抽取正确": "无法判断",
            "人工_候选记忆是否明确反驳问题前提": "是",
            "人工_是否只是相关但不反驳": "否",
            "人工_是否只是缺失信息而非反证": "否",
            "人工_是否存在时间不匹配": "否",
            "人工_是否足以支持答案": "是",
            "人工_是否应该触发R3a": "是",
            "人工_最终标签": "真反证",
            "人工_错误来源": "R3a漏触发",
            "人工_备注": "同一对象和场景；候选记忆明确说 herbal teas 已不再有效，反驳“满意/能维持能量”的问题前提。",
        }
    if row_id in AMBIGUOUS_ROWS:
        return {
            "人工_是否同主题": "是",
            "人工_问题前提是否抽取正确": "无法判断",
            "人工_候选记忆是否明确反驳问题前提": "模糊",
            "人工_是否只是相关但不反驳": "部分",
            "人工_是否只是缺失信息而非反证": "否",
            "人工_是否存在时间不匹配": "是",
            "人工_是否足以支持答案": "部分",
            "人工_是否应该触发R3a": "不确定",
            "人工_最终标签": "模糊反证",
            "人工_错误来源": "时间判断问题",
            "人工_备注": "记忆表达 continued commitment to hospitality，可反驳“完全离开行业”，但时间早于问题多年，不能算明确漏触发。",
        }
    if row_id in USEFUL_SAME:
        return {
            "人工_是否同主题": "是",
            "人工_问题前提是否抽取正确": "无法判断",
            "人工_候选记忆是否明确反驳问题前提": "否",
            "人工_是否只是相关但不反驳": "是",
            "人工_是否只是缺失信息而非反证": "部分",
            "人工_是否存在时间不匹配": "是",
            "人工_是否足以支持答案": "部分",
            "人工_是否应该触发R3a": "否",
            "人工_最终标签": "非反证但有用",
            "人工_错误来源": "候选发现器误报",
            "人工_备注": "同一对象或同一偏好/职业领域的旧状态、背景或相近信息，但没有明确推翻问题前提。",
        }
    if row_id in USEFUL_PARTIAL:
        return {
            "人工_是否同主题": "部分",
            "人工_问题前提是否抽取正确": "无法判断",
            "人工_候选记忆是否明确反驳问题前提": "否",
            "人工_是否只是相关但不反驳": "是",
            "人工_是否只是缺失信息而非反证": "部分",
            "人工_是否存在时间不匹配": "不确定",
            "人工_是否足以支持答案": "否",
            "人工_是否应该触发R3a": "否",
            "人工_最终标签": "非反证但有用",
            "人工_错误来源": "候选发现器误报",
            "人工_备注": "只共享人物、场景或宽泛领域，缺少问题中的具体对象、动作或时间；不构成反证。",
        }
    return {
        "人工_是否同主题": "部分",
        "人工_问题前提是否抽取正确": "无法判断",
        "人工_候选记忆是否明确反驳问题前提": "否",
        "人工_是否只是相关但不反驳": "部分",
        "人工_是否只是缺失信息而非反证": "是",
        "人工_是否存在时间不匹配": "不确定",
        "人工_是否足以支持答案": "否",
        "人工_是否应该触发R3a": "否",
        "人工_最终标签": "非反证",
        "人工_错误来源": "候选发现器误报",
        "人工_备注": "候选记忆仅命中人名、否定词或宽泛锚点，未涉及问题前提对应的具体属性或事件。",
    }


def rebuild_validations(wb, ws, col: dict[str, int]) -> None:
    if "_choices" in wb.sheetnames:
        choices = wb["_choices"]
        choices.delete_rows(1, choices.max_row)
    else:
        choices = wb.create_sheet("_choices")
    choices.sheet_state = "hidden"

    for cidx, field in enumerate(VALIDATION_CHOICES, start=1):
        choices.cell(row=1, column=cidx, value=field)
        for ridx, value in enumerate(VALIDATION_CHOICES[field], start=2):
            choices.cell(row=ridx, column=cidx, value=value)

    ws.data_validations.dataValidation = []
    for cidx, field in enumerate(VALIDATION_CHOICES, start=1):
        values = VALIDATION_CHOICES[field]
        excel_col = ws.cell(row=1, column=col[field]).column_letter
        choice_col = choices.cell(row=1, column=cidx).column_letter
        formula = f"'_choices'!${choice_col}$2:${choice_col}${len(values) + 1}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{excel_col}2:{excel_col}{ws.max_row}")


def main() -> None:
    root = Path.cwd()
    outdir = root / "outputs" / "r3a_v2_2_p1_manual_audit_prep"
    xlsx_path = outdir / "p1_manual_audit_sheet.xlsx"
    csv_path = outdir / "p1_manual_audit_sheet.csv"

    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}

    for row_idx in range(2, ws.max_row + 1):
        row_id = ws.cell(row_idx, col["row_id"]).value
        values = judgement(str(row_id))
        for field, value in values.items():
            ws.cell(row_idx, col[field], value)

    manual_fill = PatternFill("solid", fgColor="FFF2CC")
    for field in MANUAL_FIELDS:
        ws.cell(1, col[field]).fill = manual_fill
    rebuild_validations(wb, ws, col)
    wb.save(xlsx_path)

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
        fieldnames = list(rows[0].keys()) if rows else headers
    for row in rows:
        row.update(judgement(row["row_id"]))
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    label_counts = Counter(judgement(row["row_id"])["人工_最终标签"] for row in rows)
    should_counts = Counter(judgement(row["row_id"])["人工_是否应该触发R3a"] for row in rows)
    source_counts = Counter(judgement(row["row_id"])["人工_错误来源"] for row in rows)
    summary = {
        "round": "0016",
        "input_round": "round_0015",
        "task": "fill_p1_manual_audit_sheet_from_5shen_guidelines",
        "audit_file": str(xlsx_path),
        "candidate_rows": len(rows),
        "unique_samples": len({row["sample_id"] for row in rows}),
        "manual_final_label_counts": dict(label_counts),
        "manual_should_trigger_counts": dict(should_counts),
        "manual_failure_source_counts": dict(source_counts),
        "final_true_counterevidence_rows": label_counts.get("真反证", 0),
        "final_ambiguous_counterevidence_rows": label_counts.get("模糊反证", 0),
        "final_should_trigger_rows": should_counts.get("是", 0),
        "api_calls": 0,
        "changed_selector_logic": False,
        "changed_prompt": False,
        "changed_retriever": False,
        "changed_memory_store": False,
        "audit_principle": "Conservative: related memory is not true counterevidence unless it clearly refutes the question premise.",
    }
    (outdir / "p1_manual_audit_5shen_filled_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    report = f"""# P1 第一优先级人工审计填写结果（5审）

## 输入

- 审计规则：`D:\\幻觉\\临时\\人工审查\\5审.md`
- 审计表：`{xlsx_path}`

## 填写原则

本轮按保守原则填写：只有候选记忆明确推翻问题前提，且对象、属性、时间范围足够对应，才标记为“真反证”和“应该触发R3a”。相关、旧背景、缺失信息、对象错位和时间错位均不作为真反证。

由于表格未提供 R3a 的显式问题前提抽取结果，`人工_问题前提是否抽取正确` 统一填为“无法判断”。

## 汇总

- 候选行数：{summary["candidate_rows"]}
- 唯一样本数：{summary["unique_samples"]}
- 最终标签分布：`{dict(label_counts)}`
- 是否应该触发 R3a 分布：`{dict(should_counts)}`
- 错误来源分布：`{dict(source_counts)}`
- 真反证行数：{summary["final_true_counterevidence_rows"]}
- 模糊反证行数：{summary["final_ambiguous_counterevidence_rows"]}
- 应该触发 R3a 行数：{summary["final_should_trigger_rows"]}

## 关键说明

P1-0077 被标为真反证：候选记忆明确说明 herbal teas 在 meal-plan planning 中已不再有效，直接反驳“对 herbal teas 维持能量感到满意”的问题前提。

P1-0036 被标为模糊反证：候选记忆与 hospitality commitment 有关，但时间早于问题多年，不能直接作为明确漏触发。

其余多数候选是同名、同领域、旧偏好或宽泛背景命中，属于高召回候选发现器误报或非反证但有用信息。

## 安全确认

- API 调用：0
- 未修改 R3a selector
- 未修改 prompt
- 未修改 retriever
- 未修改 memory store
"""
    (outdir / "p1_manual_audit_5shen_filled_report.md").write_text(report, encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
