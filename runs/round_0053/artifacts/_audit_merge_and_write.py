from __future__ import annotations
import json, shutil, datetime
from pathlib import Path
from collections import Counter
import openpyxl

WORKDIR = Path(r"D:\幻觉\临时\人工审查\审查生成错误率\audit_work")
WB = Path(r"D:\幻觉\临时\人工审查\审查生成错误率\审核-多轮回答评估-G.xlsx")
SRC = Path(r"D:\幻觉\outputs\e1_memos_generation_failure_oracle_v1\case_summary.jsonl")
REVIEWER = "Kiro-AI代审"
REVIEW_DATE = datetime.date(2026, 7, 15)

SUFF_OK = {"充分", "部分", "不充分", "不确定"}
RUN_OK = {"C", "H", "O"}
CONF_OK = {"高", "中", "低"}

def final_class(suff: str, runs: list[str]) -> str:
    if suff in {"部分", "不充分", "不确定"}:
        return "evidence_definition_failure"
    if suff != "充分":
        return "unresolved"
    c = sum(1 for r in runs if r == "C")
    if c == 0:
        return "robust_generation_failure"
    if c == 3:
        return "ua3_representation_or_admission_failure"
    return "generation_instability"

# 1) merge batch outputs
labels: dict[str, dict] = {}
for i in range(1, 14):
    fp = WORKDIR / f"batch_{i:02d}_out.json"
    data = json.loads(fp.read_text(encoding="utf-8"))
    for r in data:
        cid = r["case_id"]
        assert cid not in labels, f"duplicate {cid}"
        assert r["sufficiency"] in SUFF_OK, f"bad suff {cid} {r['sufficiency']}"
        for k in ("run1", "run2", "run3"):
            assert r[k] in RUN_OK, f"bad {k} {cid} {r[k]}"
        assert r["confidence"] in CONF_OK, f"bad conf {cid} {r['confidence']}"
        labels[cid] = r
print("merged labels:", len(labels))

# 2) validate against workbook case_ids
wb = openpyxl.load_workbook(WB, data_only=False)
ws = wb["人工审核"]
rows = list(range(2, ws.max_row + 1))
wb_ids = {ws.cell(row=r, column=2).value: r for r in rows}
missing = [c for c in wb_ids if c not in labels]
extra = [c for c in labels if c not in wb_ids]
print("workbook rows:", len(wb_ids), "missing_from_labels:", missing, "extra_labels:", extra)
assert not missing and not extra, "case_id mismatch"

# 3) load machine auto classification for comparison
machine = {}
for line in SRC.read_text(encoding="utf-8").splitlines():
    if line.strip():
        o = json.loads(line)
        machine[o["case_id"]] = o.get("automatic_classification")

# 4) write into workbook (skip already-filled rows = user's example)
COL = {"L":12,"M":13,"N":14,"O":15,"P":16,"Q":17,"R":18,"S":19,"T":20,"U":21,"V":22}
written = 0
preserved = []
human_final = {}
for cid, r in labels.items():
    row = wb_ids[cid]
    runs = [r["run1"], r["run2"], r["run3"]]
    fc = final_class(r["sufficiency"], runs)
    human_final[cid] = fc
    existing_L = ws.cell(row=row, column=COL["L"]).value
    if existing_L not in (None, ""):
        preserved.append((cid, row))
        continue
    ws.cell(row=row, column=COL["L"]).value = r["sufficiency"]
    ws.cell(row=row, column=COL["M"]).value = r["run1"]
    ws.cell(row=row, column=COL["N"]).value = r["run2"]
    ws.cell(row=row, column=COL["O"]).value = r["run3"]
    # P is a formula (建议最终分类) -> leave intact
    ws.cell(row=row, column=COL["Q"]).value = fc            # 人工_最终分类
    ws.cell(row=row, column=COL["R"]).value = r["confidence"]
    ws.cell(row=row, column=COL["S"]).value = REVIEWER
    ws.cell(row=row, column=COL["T"]).value = REVIEW_DATE
    ws.cell(row=row, column=COL["U"]).value = r.get("note", "")
    # V is a formula (审核完成) -> leave intact
    written += 1

# 5) backup then save
bak = WB.with_name(WB.stem + f"_backup_{datetime.datetime.now():%Y%m%d_%H%M%S}.xlsx")
shutil.copy2(WB, bak)
wb.save(WB)
print("written rows:", written, "preserved(existing):", preserved)
print("backup:", bak.name)

# 6) summaries
fin_counts = Counter(human_final.values())
print("\n=== Human final classification (all 381, incl. row2 label) ===")
for k in ["robust_generation_failure","generation_instability","ua3_representation_or_admission_failure","evidence_definition_failure","unresolved"]:
    print(f"  {k}: {fin_counts.get(k,0)}")
suff_counts = Counter(v["sufficiency"] for v in labels.values())
print("sufficiency:", dict(suff_counts))
conf_counts = Counter(v["confidence"] for v in labels.values())
print("confidence:", dict(conf_counts))

# machine vs human final agreement
agree = sum(1 for c in labels if machine.get(c) == human_final[c])
print(f"\nmachine-vs-human final agreement: {agree}/{len(labels)} = {agree/len(labels):.1%}")
disagree = [(c, machine.get(c), human_final[c]) for c in labels if machine.get(c) != human_final[c]]
dc = Counter((m, h) for _, m, h in disagree)
print("top disagreement transitions (machine -> human):")
for (m, h), n in dc.most_common(12):
    print(f"  {m} -> {h}: {n}")

# write merged results file
out = []
for cid, r in labels.items():
    out.append({**r, "final_classification": human_final[cid], "machine_auto": machine.get(cid),
                "row": wb_ids[cid]})
out.sort(key=lambda x: x["row"])
(WORKDIR / "audit_results_merged.json").write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
print("\nmerged results ->", (WORKDIR / "audit_results_merged.json").name)
