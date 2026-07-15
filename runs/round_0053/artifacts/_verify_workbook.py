from __future__ import annotations
import json
from pathlib import Path
from collections import Counter
import openpyxl

WB = r"D:\幻觉\临时\人工审查\审查生成错误率\审核-多轮回答评估-G.xlsx"
MERGED = Path(r"D:\幻觉\临时\人工审查\审查生成错误率\audit_work\audit_results_merged.json")

merged = {r["case_id"]: r for r in json.loads(MERGED.read_text(encoding="utf-8"))}

wb = openpyxl.load_workbook(WB, data_only=False)
ws = wb["人工审核"]
COL = {"L":12,"M":13,"N":14,"O":15,"Q":17,"R":18,"S":19,"U":21}

mismatch = []
s_vals = Counter()
filled = 0
for r in range(2, ws.max_row+1):
    cid = ws.cell(row=r, column=2).value
    L = ws.cell(row=r, column=COL["L"]).value
    M = ws.cell(row=r, column=COL["M"]).value
    N = ws.cell(row=r, column=COL["N"]).value
    O = ws.cell(row=r, column=COL["O"]).value
    Q = ws.cell(row=r, column=COL["Q"]).value
    S = ws.cell(row=r, column=COL["S"]).value
    s_vals[S] += 1
    if L not in (None,""):
        filled += 1
    m = merged.get(cid)
    if not m:
        continue
    exp = (m["sufficiency"], m["run1"], m["run2"], m["run3"], m["final_classification"])
    got = (L, M, N, O, Q)
    if got != exp and r != 2:  # row2 is user's, expected to differ
        mismatch.append((r, cid, got, exp, S))

print("rows filled (L nonempty):", filled, "/", ws.max_row-1)
print("审核人(S) distribution:", dict(s_vals))
print("mismatches vs my merged labels (excl row2):", len(mismatch))
for r,cid,got,exp,S in mismatch[:20]:
    print(f"  row{r} {cid} S={S}\n    wb ={got}\n    mine={exp}")
