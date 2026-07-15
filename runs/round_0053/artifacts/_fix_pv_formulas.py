from __future__ import annotations
import openpyxl
WB = r"D:\幻觉\临时\人工审查\审查生成错误率\审核-多轮回答评估-G.xlsx"
wb = openpyxl.load_workbook(WB, data_only=False)
ws = wb["人工审核"]
for r in range(2, ws.max_row+1):
    ws.cell(row=r, column=16).value = (
        f'=IF(COUNTA(L{r}:O{r})<4,"未完成",'
        f'IF(OR(L{r}="部分",L{r}="不充分",L{r}="不确定"),"evidence_definition_failure",'
        f'IF(L{r}<>"充分","unresolved",'
        f'IF(COUNTIF(M{r}:O{r},"C")=0,"robust_generation_failure",'
        f'IF(COUNTIF(M{r}:O{r},"C")=3,"ua3_representation_or_admission_failure","generation_instability")))))'
    )
    ws.cell(row=r, column=22).value = (
        f'=IF(AND(L{r}<>"",M{r}<>"",N{r}<>"",O{r}<>"",Q{r}<>"",R{r}<>""),"完成","未完成")'
    )
wb.save(WB)
print("restored P/V formulas for rows 2..", ws.max_row)
# quick recheck
wb2 = openpyxl.load_workbook(WB, data_only=False)
ws2 = wb2["人工审核"]
for r in (2,3,382):
    print(r, "P:", str(ws2.cell(row=r,column=16).value)[:15], "V:", str(ws2.cell(row=r,column=22).value)[:15])
