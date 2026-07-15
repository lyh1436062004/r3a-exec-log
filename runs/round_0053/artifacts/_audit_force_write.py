from __future__ import annotations
import json, shutil, datetime
from pathlib import Path
from collections import Counter
import openpyxl

WB = Path(r"D:\幻觉\临时\人工审查\审查生成错误率\审核-多轮回答评估-G.xlsx")
MERGED = Path(r"D:\幻觉\临时\人工审查\审查生成错误率\audit_work\audit_results_merged.json")
REVIEWER = "Kiro-AI代审"
REVIEW_DATE = datetime.date(2026, 7, 15)
COL = {"L":12,"M":13,"N":14,"O":15,"Q":17,"R":18,"S":19,"T":20,"U":21}

merged = {r["case_id"]: r for r in json.loads(MERGED.read_text(encoding="utf-8"))}

wb = openpyxl.load_workbook(WB, data_only=False)
ws = wb["人工审核"]

# backup pristine-of-now
bak = WB.with_name(WB.stem + f"_prewrite_{datetime.datetime.now():%Y%m%d_%H%M%S}.xlsx")
shutil.copy2(WB, bak)

written = 0
skipped = []
for r in range(2, ws.max_row+1):
    cid = ws.cell(row=r, column=2).value
    if r == 2:
        skipped.append((r, cid, "用户示例保留"))
        continue
    m = merged[cid]
    ws.cell(row=r, column=COL["L"]).value = m["sufficiency"]
    ws.cell(row=r, column=COL["M"]).value = m["run1"]
    ws.cell(row=r, column=COL["N"]).value = m["run2"]
    ws.cell(row=r, column=COL["O"]).value = m["run3"]
    ws.cell(row=r, column=COL["Q"]).value = m["final_classification"]
    ws.cell(row=r, column=COL["R"]).value = m["confidence"]
    ws.cell(row=r, column=COL["S"]).value = REVIEWER
    ws.cell(row=r, column=COL["T"]).value = REVIEW_DATE
    ws.cell(row=r, column=COL["U"]).value = m.get("note", "")
    written += 1

wb.save(WB)
print("force-written:", written, "skipped:", skipped, "backup:", bak.name)
