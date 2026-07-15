from __future__ import annotations
import openpyxl
from collections import Counter
WB = r"D:\幻觉\临时\人工审查\审查生成错误率\审核-多轮回答评估-G.xlsx"
wb = openpyxl.load_workbook(WB, data_only=False)
ws = wb["人工审核"]
print("sheets:", wb.sheetnames, "dims:", ws.max_row, "x", ws.max_column)
print("P2 formula?", str(ws['P2'].value)[:20], "| V2 formula?", str(ws['V2'].value)[:20])
print("P3 formula?", str(ws['P3'].value)[:20], "| V3 formula?", str(ws['V3'].value)[:20])
print("tables:", list(ws.tables.keys()))
dvs = ws.data_validations.dataValidation
print("num data validations:", len(dvs))
for dv in dvs:
    print("  DV", dv.type, str(dv.formula1)[:40], "->", dv.sqref)
# final distributions from Q
qc = Counter(ws.cell(row=r, column=17).value for r in range(2, ws.max_row+1))
print("Q (人工_最终分类) distribution:", dict(qc))
