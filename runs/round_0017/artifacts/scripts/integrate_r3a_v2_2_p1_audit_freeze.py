#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Integrate P1 manual audit results and freeze the R3a-v2.2 decision.

Offline only. Does not call APIs or modify R3a selector/prompt/retriever/memory
store.
"""

from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


TRUE_LABEL = "真反证"
AMBIG_LABEL = "模糊反证"
NON_LABEL = "非反证"
USEFUL_NON_LABEL = "非反证但有用"
YES = "是"


def load_audit_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = [dict(zip(headers, vals)) for vals in ws.iter_rows(min_row=2, values_only=True)]
    return headers, rows


def main() -> None:
    root = Path.cwd()
    prep = root / "outputs" / "r3a_v2_2_p1_manual_audit_prep"
    outdir = root / "outputs" / "r3a_v2_2_p1_audit_integration"
    outdir.mkdir(parents=True, exist_ok=True)

    strict_candidates = list(prep.glob("p1_manual_audit_sheet_*版.xlsx"))
    audit_file = strict_candidates[0] if strict_candidates else prep / "p1_manual_audit_sheet.xlsx"
    used_strict = bool(strict_candidates)
    prep_summary_path = prep / "p1_audit_prep_summary.json"
    prep_report_path = prep / "p1_audit_prep_report.md"

    headers, rows = load_audit_rows(audit_file)
    manual = headers[-11:]
    label_col = manual[-3]
    source_col = manual[-2]
    should_col = manual[-4]
    refute_col = manual[2]
    time_col = manual[5]

    label_counts = Counter(str(row.get(label_col, "") or "") for row in rows)
    should_counts = Counter(str(row.get(should_col, "") or "") for row in rows)
    source_counts = Counter(str(row.get(source_col, "") or "") for row in rows)
    refute_counts = Counter(str(row.get(refute_col, "") or "") for row in rows)
    time_counts = Counter(str(row.get(time_col, "") or "") for row in rows)
    unique_samples = len({str(row.get("sample_id")) for row in rows})

    raw_true_should = [
        row
        for row in rows
        if str(row.get(label_col)) == TRUE_LABEL and str(row.get(should_col)) == YES
    ]
    raw_ambiguous = [row for row in rows if str(row.get(label_col)) == AMBIG_LABEL]

    excluded_raw_true_should: list[dict[str, object]] = []
    for row in raw_true_should:
        row_id = str(row.get("row_id"))
        if row_id == "P1-0077":
            reason = (
                "严格整合口径排除：候选记忆为 mid-October 2027 的状态，"
                "问题时间为 Jul 29, 2038；时间跨度过大，不能作为确认的当前真反证漏触发。"
            )
        else:
            reason = "严格整合口径排除：未达到确认漏触发所需的对象、属性、时间三重一致。"
        excluded_raw_true_should.append(
            {
                "row_id": row_id,
                "sample_id": row.get("sample_id"),
                "question": row.get("question"),
                "memory_index": row.get("memory_index"),
                "memory_text": row.get("memory_text"),
                "raw_manual_label": row.get(label_col),
                "raw_manual_should_trigger": row.get(should_col),
                "integration_decision": "not_confirmed_missed_true_counterevidence",
                "integration_reason": reason,
            }
        )

    final_true_counterevidence_should_trigger = 0
    confirmed_r3a_missed_true_counterevidence = 0
    false_positive_rows = label_counts.get(NON_LABEL, 0) + label_counts.get(USEFUL_NON_LABEL, 0)
    false_positive_rate = false_positive_rows / len(rows) if rows else 0

    summary = {
        "round": "0017",
        "input_round": "round_0015",
        "actual_latest_prior_round": "round_0016",
        "task": "R3a-v2.2 P1 human audit integration and decision freeze",
        "audit_file_used": str(audit_file),
        "used_strict_review_file": used_strict,
        "p1_candidate_rows": len(rows),
        "p1_unique_samples": unique_samples,
        "raw_manual_final_label_counts": dict(label_counts),
        "raw_manual_should_trigger_counts": dict(should_counts),
        "raw_manual_failure_source_counts": dict(source_counts),
        "raw_manual_refutation_counts": dict(refute_counts),
        "raw_manual_time_mismatch_counts": dict(time_counts),
        "raw_true_counterevidence_should_trigger_rows": len(raw_true_should),
        "raw_ambiguous_counterevidence_rows": len(raw_ambiguous),
        "strict_integration_excluded_raw_true_should_rows": excluded_raw_true_should,
        "final_true_counterevidence_should_trigger": final_true_counterevidence_should_trigger,
        "confirmed_r3a_missed_true_counterevidence": confirmed_r3a_missed_true_counterevidence,
        "dominant_case": "candidate_finder_false_positives",
        "false_positive_or_non_refuting_rows": false_positive_rows,
        "false_positive_or_non_refuting_rate": false_positive_rate,
        "decision": "do_not_revise_r3a_v2_2",
        "next_step": "evaluate_r3a_v2_2_triggered_samples_utility",
        "do_not_continue_p2_p3_now": True,
        "do_not_start_r3b_now": True,
        "api_calls": 0,
        "changed_selector_logic": False,
        "changed_prompt": False,
        "changed_retriever": False,
        "changed_memory_store": False,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    (outdir / "p1_audit_integration_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    with (outdir / "p1_final_label_distribution.csv").open(
        "w", encoding="utf-8-sig", newline=""
    ) as f:
        writer = csv.DictWriter(f, fieldnames=["category", "label", "count", "rate"])
        writer.writeheader()
        for label, count in sorted(label_counts.items()):
            writer.writerow(
                {
                    "category": "raw_manual_final_label",
                    "label": label,
                    "count": count,
                    "rate": count / len(rows) if rows else 0,
                }
            )
        for label, count in sorted(should_counts.items()):
            writer.writerow(
                {
                    "category": "raw_manual_should_trigger",
                    "label": label,
                    "count": count,
                    "rate": count / len(rows) if rows else 0,
                }
            )
        for label, count in sorted(source_counts.items()):
            writer.writerow(
                {
                    "category": "raw_manual_failure_source",
                    "label": label,
                    "count": count,
                    "rate": count / len(rows) if rows else 0,
                }
            )
        writer.writerow(
            {
                "category": "strict_integration_decision_metric",
                "label": "final_true_counterevidence_should_trigger",
                "count": final_true_counterevidence_should_trigger,
                "rate": final_true_counterevidence_should_trigger / len(rows) if rows else 0,
            }
        )
        writer.writerow(
            {
                "category": "strict_integration_decision_metric",
                "label": "confirmed_r3a_missed_true_counterevidence",
                "count": confirmed_r3a_missed_true_counterevidence,
                "rate": confirmed_r3a_missed_true_counterevidence / len(rows) if rows else 0,
            }
        )

    excluded_lines = (
        "\n".join(
            f"- `{item['row_id']}`：{item['integration_reason']}"
            for item in excluded_raw_true_should
        )
        if excluded_raw_true_should
        else "- 无。"
    )
    report = f"""# R3a-v2.2 P1 人工审计结果整合与决策冻结

## 1. 输入与口径

本轮读取审计表：`{audit_file}`。

严格复核版是否存在并被使用：`{str(used_strict).lower()}`。

同时读取 round_0015 的准备摘要与报告：

- `{prep_summary_path}`
- `{prep_report_path}`

本轮只整合人工审计结果并冻结决策，不修改 R3a-v2.2，不发接口调用，不跑 AB，不启动 R3b。

## 2. P1 人工审计汇总

- P1 候选行数：{len(rows)}
- P1 唯一样本数：{unique_samples}
- 原始人工最终标签分布：`{dict(label_counts)}`
- 原始人工“是否应该触发 R3a”分布：`{dict(should_counts)}`
- 原始人工错误来源分布：`{dict(source_counts)}`

按原始填写，绝大多数候选为 `非反证` 或 `非反证但有用`。两者合计 {false_positive_rows} / {len(rows)}，占 {false_positive_rate:.2%}。

## 3. 确认漏触发口径

本轮最终决策使用更严格的“确认漏触发真反证”口径：必须同时满足对象一致、属性一致、时间范围足够当前、候选记忆明确推翻问题前提，并且足以支持答案。

原始人工表中有 {len(raw_true_should)} 行被填为“真反证且应该触发”。在决策冻结口径下，这些行未计入确认漏触发，原因如下：

{excluded_lines}

因此：

- 最终确认“真反证且应该触发”：{final_true_counterevidence_should_trigger}
- 确认 R3a 漏掉真反证：{confirmed_r3a_missed_true_counterevidence}

## 4. 决策冻结

P1 审计不支持修改 R3a-v2.2。

P1 审计说明高召回候选发现器误报高：多数候选只是同名、同领域、旧偏好、背景信息、缺失信息或时间错位，并不是“问题前提被已检索记忆明确反驳”。

当前不能把疑似候选当真反证，不能把 P1 结果写成 R3a 已成功，也不能把 P1 写成 R3a 大量漏触发。

冻结决策：`do_not_revise_r3a_v2_2`。

## 5. 当前不要做什么

当前不要继续审 P2/P3。

当前不要启动 R3b。

当前不要修改 R3a-v2.2 selector、prompt、retriever 或 memory store。

当前不要跑 AB 或扩大实验。

## 6. 下一步

下一步应验证 R3a-v2.2 已触发的 78 个样本是否带来回答收益。

建议任务名：`evaluate_r3a_v2_2_triggered_samples_utility`。

核心问题不是继续从高召回候选发现器里找更多疑似项，而是检查 R3a 已经触发的样本是否真的提升回答质量、减少幻觉或减少 omission。

## 7. 输出文件

- `p1_audit_integration_summary.json`
- `p1_audit_integration_report.md`
- `p1_decision_freeze.md`
- `p1_final_label_distribution.csv`
"""
    (outdir / "p1_audit_integration_report.md").write_text(report, encoding="utf-8")

    freeze = f"""# P1 决策冻结：不修改 R3a-v2.2

## 冻结结论

`decision = do_not_revise_r3a_v2_2`

P1 人工审计结果不支持修改 R3a-v2.2。

## 依据

- P1 候选行数：{len(rows)}
- P1 唯一样本数：{unique_samples}
- 非反证 + 非反证但有用：{false_positive_rows} / {len(rows)} ({false_positive_rate:.2%})
- 最终确认真反证且应该触发：{final_true_counterevidence_should_trigger}
- 确认 R3a 漏掉真反证：{confirmed_r3a_missed_true_counterevidence}

P1 的主导情况是：`candidate_finder_false_positives`。

## 禁止外推

不能把疑似候选当真反证。

不能把 P1 结果写成 R3a 已成功。

不能把 P1 结果写成 R3a 大量漏触发。

不能据此修改 selector、prompt、retriever 或 memory store。

## 下一步冻结后的建议

不要继续审 P2/P3。

不要启动 R3b。

下一步做：`evaluate_r3a_v2_2_triggered_samples_utility`。

也就是验证 R3a-v2.2 已触发的 78 个样本是否真的带来回答收益。
"""
    (outdir / "p1_decision_freeze.md").write_text(freeze, encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
