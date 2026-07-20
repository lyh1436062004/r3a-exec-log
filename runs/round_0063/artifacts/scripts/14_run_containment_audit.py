from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from threading import Lock
from typing import Any, Callable

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "outputs" / "e7_memos_containment_audit_v1"
SOURCE_WORKBOOK = (
    ROOT
    / "r3a-exec-log"
    / "runs"
    / "round_0054"
    / "artifacts"
    / "最终汇总表-人工机器复核整合.xlsx"
)
SAMPLES_PATH = ROOT / "outputs" / "e1_memos_full_oracle_v2" / "samples_memos_full.jsonl"
UA3_GENERATIONS_PATH = (
    ROOT / "outputs" / "e1_memos_full_oracle_v2" / "generations" / "UA3.jsonl"
)
UA3_VERDICTS_PATH = ROOT / "outputs" / "e1_memos_full_oracle_v2" / "verdicts" / "UA3.jsonl"
ENV_PATH = ROOT / "github" / "HaluMem" / "eval" / ".env"

EXPECTED_WORKBOOK_SHA256 = "db428cc47875bb808f7b6c0ce394d095e3d9470032d0ff78a6db2e4d1a4b796d"
EXPECTED_SOURCE_COUNTS = {
    "robust_generation_failure": 133,
    "ua3_representation_or_admission_failure": 158,
    "evidence_definition_failure": 90,
}
EXPECTED_CANDIDATES = 158
JUDGE_MODEL = os.getenv("EXP7_JUDGE_MODEL", "LongCat-2.0")
JUDGE_BASE_URL = os.getenv("EXP7_JUDGE_BASE_URL", "https://api.longcat.chat/openai/v1")
JUDGE_THINKING = os.getenv("EXP7_JUDGE_THINKING", "disabled")
TEMPERATURE = 0
MAX_TOKENS = 768
TIMEOUT = 300
RETRIES = 3
WORKERS = 8
JSON_RE = re.compile(r"\{.*\}", re.DOTALL)
VALID_COVERAGE = {"contained", "partial", "missing"}
VALID_TEMPORAL = {"not_needed", "preserved", "lost_in_rendering", "missing_from_raw", "ambiguous"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Experiment 7: containment audit of the frozen 158-case pool")
    parser.add_argument("--phase", choices=("prepare", "judge", "all"), default="all")
    parser.add_argument("--source-workbook", type=Path, default=SOURCE_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=OUT_DIR)
    parser.add_argument("--workers", type=int, default=WORKERS)
    parser.add_argument("--smoke-only", action="store_true")
    parser.add_argument("--yes", action="store_true")
    return parser.parse_args()


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_text(value: str) -> str:
    return sha256_bytes(value.encode("utf-8"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def append_jsonl(path: Path, row: dict[str, Any], lock: Lock) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with lock:
        with path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
            handle.flush()


def latest_ok_by_case(path: Path) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for row in read_jsonl(path):
        if row.get("ok"):
            out[str(row.get("case_id"))] = row
    return out


def recompute_generation_failure_class(
    evidence_sufficiency: Any,
    run_labels: list[Any],
) -> str:
    if evidence_sufficiency in {"部分", "不充分", "不确定"}:
        return "evidence_definition_failure"
    if evidence_sufficiency != "充分" or len(run_labels) != 3:
        return "unresolved"
    correct = sum(str(label).strip().upper() == "C" for label in run_labels)
    if correct == 0:
        return "robust_generation_failure"
    if correct == 3:
        return "ua3_representation_or_admission_failure"
    return "generation_instability"


def load_frozen_source_rows(source_workbook: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not source_workbook.exists():
        raise FileNotFoundError(source_workbook)
    workbook_sha = sha256_bytes(source_workbook.read_bytes())
    if source_workbook.resolve() == SOURCE_WORKBOOK.resolve() and workbook_sha != EXPECTED_WORKBOOK_SHA256:
        raise RuntimeError(
            f"frozen source workbook hash changed: expected {EXPECTED_WORKBOOK_SHA256}, got {workbook_sha}"
        )

    workbook = load_workbook(source_workbook, read_only=False, data_only=True)
    if "最终汇总表" not in workbook.sheetnames:
        raise RuntimeError("source workbook is missing 最终汇总表")
    sheet = workbook["最终汇总表"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    required = {
        "审核序号",
        "case_id",
        "dataset",
        "question_type",
        "Question",
        "Gold answer",
        "Gold evidence",
        "原 UA3 回答",
        "新回答 1",
        "新回答 2",
        "新回答 3",
        "最终_证据充分性",
        "最终_Run1",
        "最终_Run2",
        "最终_Run3",
        "建议最终分类",
        "最终分类",
    }
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(f"source workbook missing columns: {missing}")

    rows: list[dict[str, Any]] = []
    recomputed_counts: Counter[str] = Counter()
    recommended_class_disagreements = 0
    final_class_disagreements = 0
    for values in sheet.iter_rows(min_row=2, values_only=True):
        record = {str(header): values[index] for index, header in enumerate(headers) if header is not None}
        final_class = recompute_generation_failure_class(
            record.get("最终_证据充分性"),
            [record.get("最终_Run1"), record.get("最终_Run2"), record.get("最终_Run3")],
        )
        record["recomputed_final_class"] = final_class
        recomputed_counts[final_class] += 1
        if record.get("建议最终分类") != final_class:
            recommended_class_disagreements += 1
        if record.get("最终分类") != final_class:
            final_class_disagreements += 1
        rows.append(record)

    if len(rows) != 381:
        raise RuntimeError(f"source row invariant failed: expected 381, got {len(rows)}")
    if dict(recomputed_counts) != EXPECTED_SOURCE_COUNTS:
        raise RuntimeError(
            f"source class invariant failed: expected {EXPECTED_SOURCE_COUNTS}, got {dict(recomputed_counts)}"
        )
    if final_class_disagreements:
        raise RuntimeError(
            f"authoritative 最终分类 disagrees with recomputed labels for {final_class_disagreements} rows"
        )
    selected = [
        row
        for row in rows
        if row["recomputed_final_class"] == "ua3_representation_or_admission_failure"
    ]
    if len(selected) != EXPECTED_CANDIDATES:
        raise RuntimeError(f"candidate invariant failed: expected {EXPECTED_CANDIDATES}, got {len(selected)}")
    case_ids = [str(row["case_id"]) for row in selected]
    if len(case_ids) != len(set(case_ids)):
        raise RuntimeError("source workbook contains duplicate selected case IDs")
    provenance = {
        "source_workbook": str(source_workbook.resolve()),
        "source_workbook_sha256": workbook_sha,
        "source_sheet": sheet.title,
        "source_rows": len(rows),
        "selection_rule": (
            "recompute from 最终_证据充分性 and 最终_Run1..3; verify 最终分类 and do not trust stale 建议最终分类"
        ),
        "recomputed_class_counts": dict(recomputed_counts),
        "stale_recommended_class_disagreements": recommended_class_disagreements,
        "final_class_disagreements": final_class_disagreements,
        "candidate_count": len(selected),
        "candidate_ids_sha256": sha256_text("\n".join(case_ids)),
    }
    return selected, provenance


def build_candidate_pool(source_workbook: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    source_rows, provenance = load_frozen_source_rows(source_workbook)
    samples = {str(row["case_id"]): row for row in read_jsonl(SAMPLES_PATH)}
    ua3_generations = latest_ok_by_case(UA3_GENERATIONS_PATH)
    ua3_verdicts = latest_ok_by_case(UA3_VERDICTS_PATH)
    candidates: list[dict[str, Any]] = []
    for source in source_rows:
        case_id = str(source["case_id"])
        if case_id not in samples or case_id not in ua3_generations or case_id not in ua3_verdicts:
            raise RuntimeError(f"case {case_id} cannot be joined to sample, UA3 generation, and verdict")
        sample = samples[case_id]
        generation = ua3_generations[case_id]
        verdict = ua3_verdicts[case_id]
        if str(generation.get("condition")) != "UA3" or str(verdict.get("condition")) != "UA3":
            raise RuntimeError(f"case {case_id} has a non-UA3 joined record")
        if str(verdict.get("judge_label")) == "correct":
            raise RuntimeError(f"case {case_id} is not an original UA3 failure")
        raw_memories = list(sample.get("raw_memories") or [])
        if not raw_memories:
            raise RuntimeError(f"case {case_id} has no frozen retriever payload")
        direct_answers = [source.get(f"新回答 {index}") for index in range(1, 4)]
        direct_labels = [source.get(f"最终_Run{index}") for index in range(1, 4)]
        if any(str(label).strip().upper() != "C" for label in direct_labels):
            raise RuntimeError(f"case {case_id} is selected but does not have three final C labels")
        candidates.append(
            {
                "case_id": case_id,
                "source_audit_index": source.get("审核序号"),
                "dataset": sample.get("dataset"),
                "qa_key": sample.get("qa_key"),
                "question_type": sample.get("question_type"),
                "question": sample.get("question"),
                "gold_answer": sample.get("gold_answer"),
                "benchmark_gold_evidence": sample.get("gold_evidence"),
                "raw_retrieved_memories": raw_memories,
                "raw_retrieved_memories_hash": sha256_text(
                    json.dumps(raw_memories, ensure_ascii=False, sort_keys=True)
                ),
                "ua3_rendered_context": generation.get("context"),
                "ua3_rendered_context_hash": generation.get("context_hash"),
                "ua3_model_answer": generation.get("model_answer"),
                "ua3_judge_label": verdict.get("judge_label"),
                "direct_gold_answers": direct_answers,
                "direct_gold_final_labels": direct_labels,
                "source_final_evidence_sufficiency": source.get("最终_证据充分性"),
                "source_recomputed_final_class": source.get("recomputed_final_class"),
            }
        )
    candidates.sort(key=lambda row: str(row["case_id"]))
    if len(candidates) != EXPECTED_CANDIDATES:
        raise RuntimeError(f"joined candidate invariant failed: {len(candidates)}")
    return candidates, provenance


def audit_prompt(case: dict[str, Any]) -> str:
    raw_payload = json.dumps(case["raw_retrieved_memories"], ensure_ascii=False, indent=2)
    gold_evidence = json.dumps(case["benchmark_gold_evidence"], ensure_ascii=False, indent=2)
    direct_answers = json.dumps(case["direct_gold_answers"], ensure_ascii=False, indent=2)
    return f"""You are an independent containment auditor for a memory-QA experiment.

This is an evaluation-only audit. The benchmark gold answer and gold evidence define what information is materially required. Do not use retrieval labels, gold memory IDs, distractor labels, or any information outside the frozen payload below.

Question:
{case['question']}

Benchmark gold answer:
{case['gold_answer']}

Benchmark gold evidence:
{gold_evidence}

Frozen raw memories returned by MemOS (complete returned objects):
{raw_payload}

Exact context rendered to UA3:
{case['ua3_rendered_context']}

UA3 answer:
{case['ua3_model_answer']}

Three direct-gold answers used to define recovery:
{direct_answers}

Make three independent judgments:
1. raw_coverage: whether factual fields anywhere in the frozen raw memory objects jointly contain every material fact needed for the gold answer.
2. rendered_coverage: whether the exact UA3 rendered context jointly contains every material fact needed for the gold answer.
3. direct_gold_recovery_valid: true only if all three direct-gold answers materially answer the question consistently with the gold answer.

Coverage labels:
- contained: every material answer component is supported or directly derivable with ordinary reasoning.
- partial: at least one material component is supported, but another required component is absent or too vague.
- missing: the required answer is not supported.

Important rules:
- Semantic equivalence is enough; exact string overlap is not required.
- A broader term is not enough when the question requires a specific person, entity, time, relation, or state.
- Treat timestamps, keys, tags, preference, reasoning, and nested metadata as part of the raw object only when their values carry relevant factual content.
- Do not infer missing facts from benchmark conventions or world knowledge.
- A fixed license sentence is not answer evidence.
- If rendered_coverage is stronger than raw_coverage, flag consistency_error=true and explain why.

Return JSON only with exactly these keys:
{{
  "raw_coverage": "contained|partial|missing",
  "rendered_coverage": "contained|partial|missing",
  "direct_gold_recovery_valid": true,
  "raw_supporting_memory_indices": [1],
  "raw_supporting_fields": ["memory_value"],
  "missing_material_facts": ["..."],
  "facts_lost_in_rendering": ["..."],
  "temporal_information_status": "not_needed|preserved|lost_in_rendering|missing_from_raw|ambiguous",
  "consistency_error": false,
  "rationale": "brief evidence-grounded explanation"
}}
"""


def derive_audit_label(
    raw_coverage: str,
    rendered_coverage: str,
    direct_gold_recovery_valid: bool,
) -> str:
    if not direct_gold_recovery_valid:
        return "D_judge_error"
    if raw_coverage != "contained":
        return "A_evidence_missing"
    if rendered_coverage != "contained":
        return "B_present_not_rendered"
    return "C_rendered_not_used"


def parse_json_response(content: str) -> dict[str, Any]:
    try:
        value = json.loads(content)
    except json.JSONDecodeError:
        match = JSON_RE.search(content)
        if not match:
            raise
        value = json.loads(match.group(0))
    if not isinstance(value, dict):
        raise ValueError("audit response is not a JSON object")
    return value


def validate_audit_response(parsed: dict[str, Any]) -> dict[str, Any]:
    raw_coverage = str(parsed.get("raw_coverage", "")).strip().lower()
    rendered_coverage = str(parsed.get("rendered_coverage", "")).strip().lower()
    if raw_coverage not in VALID_COVERAGE:
        raise ValueError(f"invalid raw_coverage: {raw_coverage}")
    if rendered_coverage not in VALID_COVERAGE:
        raise ValueError(f"invalid rendered_coverage: {rendered_coverage}")
    recovery_valid = parsed.get("direct_gold_recovery_valid")
    if not isinstance(recovery_valid, bool):
        raise ValueError("direct_gold_recovery_valid must be boolean")
    temporal = str(parsed.get("temporal_information_status", "")).strip().lower()
    if temporal not in VALID_TEMPORAL:
        raise ValueError(f"invalid temporal_information_status: {temporal}")
    judge_reported_consistency_error = parsed.get("consistency_error")
    if not isinstance(judge_reported_consistency_error, bool):
        raise ValueError("consistency_error must be boolean")
    rank = {"missing": 0, "partial": 1, "contained": 2}
    implied_inconsistency = rank[rendered_coverage] > rank[raw_coverage]
    return {
        "raw_coverage": raw_coverage,
        "rendered_coverage": rendered_coverage,
        "direct_gold_recovery_valid": recovery_valid,
        "raw_supporting_memory_indices": list(parsed.get("raw_supporting_memory_indices") or []),
        "raw_supporting_fields": [str(value) for value in parsed.get("raw_supporting_fields") or []],
        "missing_material_facts": [str(value) for value in parsed.get("missing_material_facts") or []],
        "facts_lost_in_rendering": [str(value) for value in parsed.get("facts_lost_in_rendering") or []],
        "temporal_information_status": temporal,
        "judge_reported_consistency_error": judge_reported_consistency_error,
        "consistency_error": implied_inconsistency,
        "rationale": str(parsed.get("rationale", "")),
        "audit_label": derive_audit_label(raw_coverage, rendered_coverage, recovery_valid),
    }


def load_env() -> None:
    if ENV_PATH.exists():
        load_dotenv(ENV_PATH, override=True)


def make_client() -> OpenAI:
    load_env()
    api_key = os.getenv("EXP7_JUDGE_API_KEY") or os.getenv("GF_JUDGE_API_KEY")
    if not api_key:
        raise RuntimeError("EXP7_JUDGE_API_KEY or GF_JUDGE_API_KEY is not configured")
    return OpenAI(api_key=api_key, base_url=JUDGE_BASE_URL)


def model_matches(requested: str, actual: str | None) -> bool:
    if not actual:
        return False
    req = requested.strip().lower()
    got = actual.strip().lower()
    return got == req or got.startswith(req + "-") or got.endswith("/" + req)


def retry_call(label: str, fn: Callable[[], Any]) -> Any:
    last_error: Exception | None = None
    for attempt in range(1, RETRIES + 1):
        try:
            return fn()
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            status_code = getattr(exc, "status_code", None)
            if status_code is not None and status_code != 429 and status_code < 500:
                raise
            if attempt < RETRIES:
                time.sleep(min(2**attempt, 12))
    raise RuntimeError(f"{label} failed after {RETRIES} attempts: {last_error}")


def judge_request(client: OpenAI, prompt: str, max_tokens: int = MAX_TOKENS) -> Any:
    return client.chat.completions.create(
        model=JUDGE_MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=TEMPERATURE,
        response_format={"type": "json_object"},
        extra_body={"thinking": {"type": JUDGE_THINKING}},
        max_tokens=max_tokens,
        timeout=TIMEOUT,
    )


def smoke_test(client: OpenAI, out_dir: Path) -> dict[str, Any]:
    started = time.time()
    prompt = (
        'Return JSON only: {"raw_coverage":"contained","rendered_coverage":"contained",'
        '"direct_gold_recovery_valid":true,"raw_supporting_memory_indices":[],"raw_supporting_fields":[],'
        '"missing_material_facts":[],"facts_lost_in_rendering":[],"temporal_information_status":"not_needed",'
        '"consistency_error":false,"rationale":"ok"}'
    )
    response = retry_call("experiment 7 smoke", lambda: judge_request(client, prompt, 128))
    response_model = getattr(response, "model", None)
    if not model_matches(JUDGE_MODEL, response_model):
        raise RuntimeError(f"judge model mismatch: requested={JUDGE_MODEL}, response={response_model}")
    parsed = validate_audit_response(parse_json_response(response.choices[0].message.content or ""))
    result = {
        "ok": True,
        "requested_model": JUDGE_MODEL,
        "response_model": response_model,
        "base_url": JUDGE_BASE_URL,
        "thinking": JUDGE_THINKING,
        "temperature": TEMPERATURE,
        "parsed": parsed,
        "usage": response.usage.model_dump() if getattr(response, "usage", None) else {},
        "latency_ms": round((time.time() - started) * 1000, 2),
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
    }
    write_json(out_dir / "smoke_test.json", result)
    return result


def write_preparation_outputs(
    candidates: list[dict[str, Any]],
    provenance: dict[str, Any],
    out_dir: Path,
) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)
    write_jsonl(out_dir / "candidate_pool.jsonl", candidates)
    prompts = [audit_prompt(case) for case in candidates]
    prompt_template_hash = sha256_text(audit_prompt({
        "question": "{question}",
        "gold_answer": "{gold_answer}",
        "benchmark_gold_evidence": ["{gold_evidence}"],
        "raw_retrieved_memories": [{"memory_value": "{raw_memory}"}],
        "ua3_rendered_context": "{ua3_context}",
        "ua3_model_answer": "{ua3_answer}",
        "direct_gold_answers": ["{answer1}", "{answer2}", "{answer3}"],
    }))
    estimated_input_chars = sum(len(prompt) for prompt in prompts)
    manifest = {
        **provenance,
        "candidate_ids": [case["case_id"] for case in candidates],
        "candidate_pool_sha256": sha256_text(
            "\n".join(json.dumps(case, ensure_ascii=False, sort_keys=True) for case in candidates)
        ),
        "no_inference_features": [
            "gold_memory_ids",
            "partial_memory_ids",
            "other_memory_ids",
            "HaluMem distractor labels",
        ],
    }
    write_json(out_dir / "candidate_manifest.json", manifest)
    config = {
        "experiment_id": "e7_memos_containment_audit_v1",
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "hypothesis": (
            "The 158 direct-gold recoveries mix upstream content loss, adapter rendering loss, "
            "and evidence-use failure; only cases whose required facts exist in the frozen MemOS payload "
            "are repairable by the post-retrieval layer."
        ),
        "primary_metric": "counts and shares of deterministic labels A/B/C/D among the frozen 158 cases",
        "independent_variable": "evidence view: raw returned objects versus exact UA3 rendered context",
        "fixed_conditions": {
            "candidate_pool_sha256": manifest["candidate_ids_sha256"],
            "judge_model": JUDGE_MODEL,
            "thinking": JUDGE_THINKING,
            "temperature": TEMPERATURE,
            "max_tokens": MAX_TOKENS,
            "prompt_template_sha256": prompt_template_hash,
            "retriever": "frozen; no new retrieval",
            "memory_store": "frozen; no writes or rewrites",
        },
        "auditor_protocol_note": os.getenv(
            "EXP7_PROTOCOL_NOTE",
            "Primary configuration uses the explicitly selected judge model; human validation remains mandatory.",
        ),
        "experiment_matrix": [
            {
                "run_id": "E7-AUTO-ALL",
                "factor": "evidence view",
                "value": "raw payload + exact UA3 rendering",
                "n": EXPECTED_CANDIDATES,
                "expected_outcome": "separate A/B/C/D failure ownership",
            },
            {
                "run_id": "E7-HUMAN-A",
                "factor": "validation stratum",
                "value": "up to 30 deterministic raw partial/missing cases; census if fewer",
                "n": "min(30, stratum size)",
                "expected_outcome": "estimate false missing judgments",
            },
            {
                "run_id": "E7-HUMAN-CONTAINED",
                "factor": "validation stratum",
                "value": "30 deterministic samples from raw contained",
                "n": "min(30, stratum size)",
                "expected_outcome": "estimate false containment judgments",
            },
        ],
        "resource_estimate": {
            "gpu_hours": 0,
            "api_calls_including_smoke": EXPECTED_CANDIDATES + 1,
            "estimated_input_characters": estimated_input_chars,
            "estimated_input_tokens_rough": round(estimated_input_chars / 4),
            "maximum_output_tokens": EXPECTED_CANDIDATES * MAX_TOKENS,
            "monetary_cost": (
                "not hard-coded; compute from recorded token usage and the provider account's current rates"
            ),
            "expected_storage_mb": round(
                sum(len(json.dumps(case, ensure_ascii=False)) for case in candidates) * 3 / 1_000_000,
                2,
            ),
        },
        "analysis_plan": {
            "primary": "A/B/C/D counts, proportions, and Wilson 95% intervals",
            "secondary": "raw/rendered coverage, temporal-loss status, dataset and question-type strata",
            "human_validation": "30 raw missing/partial + 30 raw contained, deterministic seed 20260720",
            "failed_calls": "remain unresolved and block final analysis unless --allow-incomplete is used",
        },
        "secrets_persisted": False,
    }
    write_json(out_dir / "run_config.json", config)
    matrix_lines = [
        "# Experiment 7 Matrix and Resource Plan",
        "",
        "| Run ID | Factor | Value | Fixed config | Expected outcome |",
        "|---|---|---|---|---|",
    ]
    for row in config["experiment_matrix"]:
        matrix_lines.append(
            f"| {row['run_id']} | {row['factor']} | {row['value']} | frozen 158 pool; seed 20260720 | {row['expected_outcome']} |"
        )
    estimate = config["resource_estimate"]
    matrix_lines.extend(
        [
            "",
            "## Resource estimate",
            "",
            f"- GPU: {estimate['gpu_hours']} hours",
            f"- API calls: {estimate['api_calls_including_smoke']}",
            f"- Rough input tokens: {estimate['estimated_input_tokens_rough']:,}",
            f"- Maximum output tokens: {estimate['maximum_output_tokens']:,}",
            f"- Expected storage: {estimate['expected_storage_mb']:.2f} MB",
            f"- Cost: {estimate['monetary_cost']}",
            "",
            "## Execution",
            "",
            "```powershell",
            '& "D:\\conda_envs\\o1\\python.exe" "脚本\\3.0\\14_run_containment_audit.py" --phase all --yes',
            '& "D:\\conda_envs\\o1\\python.exe" "脚本\\3.0\\15_analyze_containment_audit.py"',
            "```",
            "",
            "## Analysis",
            "",
            "Primary output is the A/B/C/D census. Secondary analyses stratify by dataset, question type, raw/rendered coverage, and temporal-information loss. Human validation uses two fixed 30-case strata.",
        ]
    )
    (out_dir / "experiment_matrix.md").write_text("\n".join(matrix_lines) + "\n", encoding="utf-8")
    return config


def run_audit(client: OpenAI, candidates: list[dict[str, Any]], out_dir: Path, workers: int) -> None:
    path = out_dir / "audit_results.jsonl"
    existing = latest_ok_by_case(path)
    tasks: list[tuple[dict[str, Any], str, str]] = []
    for case in candidates:
        prompt = audit_prompt(case)
        prompt_hash = sha256_text(prompt)
        cache_key = sha256_text(
            "|".join([str(case["case_id"]), JUDGE_MODEL, JUDGE_THINKING, str(TEMPERATURE), prompt_hash])
        )
        prior = existing.get(str(case["case_id"]))
        if prior and prior.get("cache_key") == cache_key:
            continue
        tasks.append((case, prompt, cache_key))
    print(f"pending containment audit tasks: {len(tasks)}")
    lock = Lock()

    def worker(case: dict[str, Any], prompt: str, cache_key: str) -> dict[str, Any]:
        started = time.time()
        try:
            response = retry_call("containment audit", lambda: judge_request(client, prompt))
            response_model = getattr(response, "model", None)
            if not model_matches(JUDGE_MODEL, response_model):
                raise RuntimeError(f"judge model mismatch: {response_model}")
            json_repair = False
            try:
                parsed = validate_audit_response(
                    parse_json_response(response.choices[0].message.content or "")
                )
            except json.JSONDecodeError:
                json_repair = True
                response = retry_call(
                    "containment audit JSON repair",
                    lambda: judge_request(client, prompt, MAX_TOKENS * 2),
                )
                response_model = getattr(response, "model", None)
                if not model_matches(JUDGE_MODEL, response_model):
                    raise RuntimeError(f"JSON-repair judge model mismatch: {response_model}")
                parsed = validate_audit_response(
                    parse_json_response(response.choices[0].message.content or "")
                )
            return {
                "ok": True,
                "cache_key": cache_key,
                "case_id": case["case_id"],
                "dataset": case["dataset"],
                "question_type": case["question_type"],
                **parsed,
                "prompt_hash": sha256_text(prompt),
                "requested_model": JUDGE_MODEL,
                "response_model": response_model,
                "thinking": JUDGE_THINKING,
                "temperature": TEMPERATURE,
                "max_tokens": MAX_TOKENS,
                "json_repair": json_repair,
                "usage": response.usage.model_dump() if getattr(response, "usage", None) else {},
                "latency_ms": round((time.time() - started) * 1000, 2),
                "created_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
            }
        except Exception as exc:  # noqa: BLE001
            return {
                "ok": False,
                "cache_key": cache_key,
                "case_id": case["case_id"],
                "error_type": type(exc).__name__,
                "error": str(exc),
                "created_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
            }

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(worker, case, prompt, cache_key) for case, prompt, cache_key in tasks]
        for index, future in enumerate(as_completed(futures), 1):
            append_jsonl(path, future.result(), lock)
            if index % 25 == 0 or index == len(futures):
                print(f"audited {index}/{len(futures)}")


def write_run_counts(out_dir: Path) -> dict[str, Any]:
    rows = read_jsonl(out_dir / "audit_results.jsonl")
    latest = latest_ok_by_case(out_dir / "audit_results.jsonl")
    failures = [row for row in rows if not row.get("ok")]
    write_jsonl(out_dir / "failures.jsonl", failures)
    counts = {
        "rows": len(rows),
        "successful_unique_cases": len(latest),
        "failed_rows": len(failures),
        "expected_cases": EXPECTED_CANDIDATES,
    }
    write_json(out_dir / "run_counts.json", counts)
    return counts


def main() -> None:
    args = parse_args()
    source_workbook = args.source_workbook.resolve()
    out_dir = args.output_dir.resolve()
    candidates, provenance = build_candidate_pool(source_workbook)
    config = write_preparation_outputs(candidates, provenance, out_dir)
    print(json.dumps({"prepared": len(candidates), "provenance": provenance}, ensure_ascii=False))
    if args.phase == "prepare":
        return
    client = make_client()
    smoke = smoke_test(client, out_dir)
    config["smoke_test"] = smoke
    write_json(out_dir / "run_config.json", config)
    print(json.dumps(smoke, ensure_ascii=False))
    if args.smoke_only:
        return
    if not args.yes:
        raise RuntimeError(f"Pass --yes to acknowledge {EXPECTED_CANDIDATES} planned containment-audit API calls")
    run_audit(client, candidates, out_dir, args.workers)
    counts = write_run_counts(out_dir)
    if counts["successful_unique_cases"] != EXPECTED_CANDIDATES:
        raise RuntimeError(f"incomplete containment audit: {counts}")
    print(json.dumps(counts, ensure_ascii=False))


if __name__ == "__main__":
    main()
