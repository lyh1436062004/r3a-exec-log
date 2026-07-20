from __future__ import annotations

import argparse
import csv
import json
import math
import random
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "outputs" / "e7_memos_containment_audit_v1"
EXPECTED_CASES = 158
SEED = 20260720
LABEL_ORDER = [
    "A_evidence_missing",
    "B_present_not_rendered",
    "C_rendered_not_used",
    "D_judge_error",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze Experiment 7 containment audit")
    parser.add_argument("--output-dir", type=Path, default=OUT_DIR)
    parser.add_argument("--allow-incomplete", action="store_true")
    return parser.parse_args()


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


def write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fields} for row in rows)


def latest_ok_by_case(path: Path) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for row in read_jsonl(path):
        if row.get("ok"):
            out[str(row.get("case_id"))] = row
    return out


def wilson(k: int, n: int, z: float = 1.959963984540054) -> tuple[float, float]:
    if n == 0:
        return 0.0, 0.0
    p = k / n
    denominator = 1 + z * z / n
    center = (p + z * z / (2 * n)) / denominator
    half = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / denominator
    return max(0.0, center - half), min(1.0, center + half)


def make_case_rows(
    candidates: list[dict[str, Any]],
    results: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    coverage_rank = {"missing": 0, "partial": 1, "contained": 2}
    for candidate in candidates:
        case_id = str(candidate["case_id"])
        audit = results.get(case_id)
        raw_coverage = audit.get("raw_coverage") if audit else ""
        rendered_coverage = audit.get("rendered_coverage") if audit else ""
        implied_consistency_error = (
            raw_coverage in coverage_rank
            and rendered_coverage in coverage_rank
            and coverage_rank[rendered_coverage] > coverage_rank[raw_coverage]
        )
        row = {
            "case_id": case_id,
            "dataset": candidate.get("dataset"),
            "question_type": candidate.get("question_type"),
            "question": candidate.get("question"),
            "gold_answer": candidate.get("gold_answer"),
            "gold_evidence": json.dumps(candidate.get("benchmark_gold_evidence"), ensure_ascii=False),
            "ua3_answer": candidate.get("ua3_model_answer"),
            "direct_gold_answers": json.dumps(candidate.get("direct_gold_answers"), ensure_ascii=False),
            "raw_retrieved_memories": json.dumps(candidate.get("raw_retrieved_memories"), ensure_ascii=False),
            "ua3_rendered_context": candidate.get("ua3_rendered_context"),
            "raw_coverage": raw_coverage,
            "rendered_coverage": rendered_coverage,
            "direct_gold_recovery_valid": audit.get("direct_gold_recovery_valid") if audit else "",
            "audit_label": audit.get("audit_label") if audit else "unresolved",
            "raw_supporting_memory_indices": json.dumps(
                audit.get("raw_supporting_memory_indices") if audit else [], ensure_ascii=False
            ),
            "raw_supporting_fields": json.dumps(
                audit.get("raw_supporting_fields") if audit else [], ensure_ascii=False
            ),
            "missing_material_facts": json.dumps(
                audit.get("missing_material_facts") if audit else [], ensure_ascii=False
            ),
            "facts_lost_in_rendering": json.dumps(
                audit.get("facts_lost_in_rendering") if audit else [], ensure_ascii=False
            ),
            "temporal_information_status": audit.get("temporal_information_status") if audit else "",
            "judge_reported_consistency_error": audit.get("consistency_error") if audit else "",
            "consistency_error": implied_consistency_error,
            "rationale": audit.get("rationale") if audit else "",
            "judge_model": audit.get("response_model") if audit else "",
            "prompt_hash": audit.get("prompt_hash") if audit else "",
        }
        rows.append(row)
    return rows


def deterministic_human_sample(rows: list[dict[str, Any]], n_per_stratum: int = 30) -> list[dict[str, Any]]:
    rng = random.Random(SEED)
    missing = sorted(
        [row for row in rows if row["raw_coverage"] in {"partial", "missing"}],
        key=lambda row: row["case_id"],
    )
    contained = sorted(
        [row for row in rows if row["raw_coverage"] == "contained"],
        key=lambda row: row["case_id"],
    )
    rng.shuffle(missing)
    rng.shuffle(contained)
    selected: list[dict[str, Any]] = []
    for stratum, pool in (("raw_partial_or_missing", missing), ("raw_contained", contained)):
        for row in pool[: min(n_per_stratum, len(pool))]:
            selected.append({**row, "validation_stratum": stratum})
    selected_ids = {row["case_id"] for row in selected}
    mandatory = [
        row
        for row in rows
        if row["audit_label"] == "D_judge_error" or bool(row["consistency_error"])
    ]
    for row in mandatory:
        if row["case_id"] not in selected_ids:
            selected.append({**row, "validation_stratum": "mandatory_anomaly"})
            selected_ids.add(row["case_id"])
    selected.sort(key=lambda row: (row["validation_stratum"], row["case_id"]))
    return selected


def write_human_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Containment人工复核"
    fields = [
        "validation_stratum",
        "case_id",
        "dataset",
        "question_type",
        "question",
        "gold_answer",
        "gold_evidence",
        "raw_retrieved_memories",
        "ua3_rendered_context",
        "ua3_answer",
        "direct_gold_answers",
        "raw_coverage",
        "rendered_coverage",
        "audit_label",
        "temporal_information_status",
        "rationale",
        "human_raw_coverage",
        "human_rendered_coverage",
        "human_final_label",
        "human_confidence",
        "human_notes",
    ]
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 24,
        "B": 22,
        "C": 12,
        "D": 28,
        "E": 45,
        "F": 35,
        "G": 55,
        "H": 70,
        "I": 65,
        "J": 35,
        "K": 45,
        "L": 16,
        "M": 18,
        "N": 28,
        "O": 24,
        "P": 55,
        "Q": 18,
        "R": 22,
        "S": 30,
        "T": 18,
        "U": 45,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    coverage_validation = DataValidation(type="list", formula1='"contained,partial,missing,uncertain"')
    label_validation = DataValidation(
        type="list",
        formula1='"A_evidence_missing,B_present_not_rendered,C_rendered_not_used,D_judge_error,uncertain"',
    )
    confidence_validation = DataValidation(type="list", formula1='"high,medium,low"')
    sheet.add_data_validation(coverage_validation)
    sheet.add_data_validation(label_validation)
    sheet.add_data_validation(confidence_validation)
    if sheet.max_row >= 2:
        coverage_validation.add(f"Q2:R{sheet.max_row}")
        label_validation.add(f"S2:S{sheet.max_row}")
        confidence_validation.add(f"T2:T{sheet.max_row}")
    guide = workbook.create_sheet("标签说明")
    guide_rows = [
        ("标签", "定义", "责任边界"),
        ("A_evidence_missing", "冻结 raw memories 缺少完整关键事实", "memory writing/compression/retrieval content"),
        ("B_present_not_rendered", "raw 对象含完整事实，但 UA3 context 未完整呈现", "unified adapter"),
        ("C_rendered_not_used", "UA3 context 已含完整事实，但 UA3 回答仍错", "decision/rendering/prompt/generation"),
        ("D_judge_error", "实验6的恢复判断不成立", "evaluation judge"),
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
    guide.column_dimensions["A"].width = 30
    guide.column_dimensions["B"].width = 70
    guide.column_dimensions["C"].width = 45
    workbook.save(path)


def sum_usage(results: list[dict[str, Any]]) -> dict[str, int]:
    totals: Counter[str] = Counter()
    for row in results:
        for key, value in (row.get("usage") or {}).items():
            if isinstance(value, int):
                totals[key] += value
    return dict(totals)


def main() -> None:
    args = parse_args()
    out_dir = args.output_dir.resolve()
    candidates = read_jsonl(out_dir / "candidate_pool.jsonl")
    results_map = latest_ok_by_case(out_dir / "audit_results.jsonl")
    if len(candidates) != EXPECTED_CASES:
        raise RuntimeError(f"candidate pool invariant failed: expected {EXPECTED_CASES}, got {len(candidates)}")
    if len(results_map) != EXPECTED_CASES and not args.allow_incomplete:
        raise RuntimeError(
            f"incomplete experiment: expected {EXPECTED_CASES} successful audits, got {len(results_map)}"
        )
    rows = make_case_rows(candidates, results_map)
    label_counts = Counter(str(row["audit_label"]) for row in rows)
    raw_counts = Counter(str(row["raw_coverage"]) for row in rows)
    rendered_counts = Counter(str(row["rendered_coverage"]) for row in rows)
    temporal_counts = Counter(str(row["temporal_information_status"]) for row in rows)

    label_statistics: list[dict[str, Any]] = []
    for label in LABEL_ORDER + (["unresolved"] if label_counts["unresolved"] else []):
        count = label_counts[label]
        low, high = wilson(count, len(rows))
        label_statistics.append(
            {
                "label": label,
                "count": count,
                "share": count / len(rows),
                "wilson_95_low": low,
                "wilson_95_high": high,
            }
        )

    grouped: list[dict[str, Any]] = []
    for field in ("dataset", "question_type"):
        groups: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            groups[str(row[field])].append(row)
        for value, group_rows in sorted(groups.items()):
            counts = Counter(str(row["audit_label"]) for row in group_rows)
            grouped.append(
                {
                    "field": field,
                    "value": value,
                    "n": len(group_rows),
                    **{label: counts[label] for label in LABEL_ORDER},
                }
            )

    validation_rows = deterministic_human_sample(rows)
    fields = list(rows[0].keys()) if rows else []
    validation_fields = ["validation_stratum", *fields, "human_raw_coverage", "human_rendered_coverage", "human_final_label", "human_confidence", "human_notes"]
    for row in validation_rows:
        row.update(
            {
                "human_raw_coverage": "",
                "human_rendered_coverage": "",
                "human_final_label": "",
                "human_confidence": "",
                "human_notes": "",
            }
        )
    write_jsonl(out_dir / "containment_case_summary.jsonl", rows)
    write_csv(out_dir / "containment_case_summary.csv", rows, fields)
    write_jsonl(out_dir / "human_validation_sample.jsonl", validation_rows)
    write_csv(out_dir / "human_validation_sample.csv", validation_rows, validation_fields)
    write_human_workbook(out_dir / "human_validation_sample.xlsx", validation_rows)

    result_rows = list(results_map.values())
    result = {
        "experiment_id": "e7_memos_containment_audit_v1",
        "coverage": {
            "candidates": len(candidates),
            "successful_audits": len(results_map),
            "unresolved": label_counts["unresolved"],
        },
        "label_counts": dict(label_counts),
        "label_statistics": label_statistics,
        "raw_coverage_counts": dict(raw_counts),
        "rendered_coverage_counts": dict(rendered_counts),
        "temporal_information_status_counts": dict(temporal_counts),
        "post_retrieval_repairable": {
            "count_B_plus_C": label_counts["B_present_not_rendered"] + label_counts["C_rendered_not_used"],
            "share_of_158": (
                label_counts["B_present_not_rendered"] + label_counts["C_rendered_not_used"]
            )
            / len(rows),
            "adapter_specific_B": label_counts["B_present_not_rendered"],
            "evidence_use_specific_C": label_counts["C_rendered_not_used"],
        },
        "grouped": grouped,
        "judge_usage": sum_usage(result_rows),
        "human_validation": {
            "seed": SEED,
            "requested_per_stratum": 30,
            "sample_rows": len(validation_rows),
            "strata": dict(Counter(row["validation_stratum"] for row in validation_rows)),
            "status": "prepared; requires human labels",
        },
    }
    write_json(out_dir / "results.json", result)

    lines = [
        "# Experiment 7 - Containment Audit",
        "",
        "## Coverage",
        "",
        f"- Frozen candidates: {len(candidates)}/{EXPECTED_CASES}",
        f"- Successful automatic audits: {len(results_map)}/{EXPECTED_CASES}",
        "- Retriever and memory store: frozen; no retrieval or write was performed",
        "- Inference-feature exclusion: no gold memory IDs, partial/distractor labels, or HaluMem distractor labels were supplied to the auditor",
        "",
        "## Primary A/B/C/D census",
        "",
        "| Label | Count | Share | Wilson 95% CI |",
        "|---|---:|---:|---:|",
    ]
    for stat in label_statistics:
        lines.append(
            f"| {stat['label']} | {stat['count']} | {stat['share']:.2%} | [{stat['wilson_95_low']:.2%}, {stat['wilson_95_high']:.2%}] |"
        )
    repairable = result["post_retrieval_repairable"]
    lines.extend(
        [
            "",
            "## Interpretation",
            "",
            f"- Upstream content loss (A): {label_counts['A_evidence_missing']}/{len(rows)}",
            f"- Unified-adapter rendering headroom (B): {label_counts['B_present_not_rendered']}/{len(rows)}",
            f"- Already rendered but not used (C): {label_counts['C_rendered_not_used']}/{len(rows)}",
            f"- Direct-gold recovery/judge error (D): {label_counts['D_judge_error']}/{len(rows)}",
            f"- Frozen-retrieval post-retrieval repairable space (B+C): {repairable['count_B_plus_C']}/{len(rows)} ({repairable['share_of_158']:.2%})",
            "",
            "## Coverage diagnostics",
            "",
            f"- Raw coverage: {dict(raw_counts)}",
            f"- Rendered coverage: {dict(rendered_counts)}",
            f"- Temporal information: {dict(temporal_counts)}",
            "",
            "## Resource usage",
            "",
            f"- GPU hours: 0",
            f"- Automatic audit calls: {len(results_map)}",
            f"- Recorded token usage: {result['judge_usage']}",
            "- Monetary cost is not asserted because provider pricing was not frozen in the experiment config; token counts are preserved for billing reconstruction.",
            "",
            "## Human validation",
            "",
            (
                f"A deterministic validation workbook with {len(validation_rows)} rows was prepared: "
                f"{result['human_validation']['strata'].get('raw_partial_or_missing', 0)} raw partial/missing "
                f"and {result['human_validation']['strata'].get('raw_contained', 0)} raw contained cases, "
                f"plus {result['human_validation']['strata'].get('mandatory_anomaly', 0)} mandatory anomalies "
                "not already sampled. The partial/missing stratum is a census when it contains fewer than 30 cases."
            ),
        ]
    )
    (out_dir / "results.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("\n".join(lines))


if __name__ == "__main__":
    main()
