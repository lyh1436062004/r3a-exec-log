#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Prepare P1 manual-audit sheets and verify round_0014 label distribution scope.

This script is offline only. It does not call APIs, generate answers, or modify
R3a selector/prompt/retriever/memory-store code.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROUND = "0015"
INPUT_ROUND = "round_0014"
P1_GROUP = "P1_memory_conflict_omission_untriggered_suspected"
EXPECTED_P1_ROWS = 120

AUTO_FIELDS = [
    "row_id",
    "sample_id",
    "qa_key",
    "question",
    "question_type",
    "r0_label",
    "baseline_response",
    "gold_answer",
    "evidence",
    "memory_index",
    "memory_text",
    "suspected_reason",
    "candidate_score",
    "r3a_triggered",
    "r3a_reject_reason",
    "gate_trace_summary",
    "raw_memory_count",
    "candidate_rank_within_sample",
    "same_sample_candidate_count",
]

MANUAL_FIELDS = [
    "人工_是否同主题",
    "人工_问题前提是否抽取正确",
    "人工_候选记忆是否明确反驳问题前提",
    "人工_是否只是相关但不反驳",
    "人工_是否只是缺失信息而非反证",
    "人工_是否存在时间不匹配",
    "人工_是否足以支持答案",
    "人工_是否应该触发R3a",
    "人工_最终标签",
    "人工_错误来源",
    "人工_备注",
]

FINAL_LABEL_CHOICES = [
    "真反证",
    "模糊反证",
    "非反证但有用",
    "非反证",
    "无法判断",
]

FAILURE_SOURCE_CHOICES = [
    "R3a漏触发",
    "候选发现器误报",
    "问题前提抽取错误",
    "槽位归类错误",
    "时间判断问题",
    "明确反驳判断问题",
    "检索证据不足",
    "不属于R3a",
    "无法判断",
]

YES_NO_UNSURE_CHOICES = ["是", "否", "不确定"]


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_jsonl_by_idx(path: Path) -> dict[int, dict[str, Any]]:
    records: dict[int, dict[str, Any]] = {}
    with path.open("r", encoding="utf-8") as f:
        for idx, line in enumerate(f):
            if not line.strip():
                continue
            obj = json.loads(line)
            sample_idx = int(obj.get("idx", idx))
            records[sample_idx] = obj
    return records


def load_qa_by_idx(path: Path) -> dict[int, dict[str, Any]]:
    records: dict[int, dict[str, Any]] = {}
    with path.open("r", encoding="utf-8") as f:
        for idx, line in enumerate(f):
            if not line.strip():
                continue
            obj = json.loads(line)
            records[idx] = obj
    return records


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def to_int(value: Any, default: int = 0) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return default


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(str(value).strip())
    except Exception:
        return default


def compact_json(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def distribution_from_sample_census(
    sample_census: dict[int, dict[str, Any]],
    *,
    memory_conflict_only: bool,
) -> Counter[str]:
    counts: Counter[str] = Counter()
    for obj in sample_census.values():
        if memory_conflict_only and obj.get("question_type") != "Memory Conflict":
            continue
        if normalize_bool(obj.get("r3a_triggered")):
            continue
        if not normalize_bool(obj.get("has_suspected_counterevidence")):
            continue
        counts[str(obj.get("baseline_label", "")).strip()] += 1
    return counts


def distribution_from_untriggered_csv(rows: list[dict[str, str]]) -> Counter[str]:
    counts: Counter[str] = Counter()
    for row in rows:
        counts[str(row.get("baseline_label", "")).strip()] += 1
    return counts


def plain_counter(counter: Counter[str]) -> dict[str, int]:
    return {k: int(counter.get(k, 0)) for k in sorted(counter.keys()) if k}


def add_xlsx_validation(wb: Workbook, ws, data_rows: int) -> None:
    choices = wb.create_sheet("_choices")
    choices.sheet_state = "hidden"

    for idx, value in enumerate(YES_NO_UNSURE_CHOICES, start=1):
        choices.cell(row=idx, column=1, value=value)
    for idx, value in enumerate(FINAL_LABEL_CHOICES, start=1):
        choices.cell(row=idx, column=2, value=value)
    for idx, value in enumerate(FAILURE_SOURCE_CHOICES, start=1):
        choices.cell(row=idx, column=3, value=value)

    yes_no_formula = f"='_choices'!$A$1:$A${len(YES_NO_UNSURE_CHOICES)}"
    label_formula = f"='_choices'!$B$1:$B${len(FINAL_LABEL_CHOICES)}"
    failure_formula = f"='_choices'!$C$1:$C${len(FAILURE_SOURCE_CHOICES)}"

    header = [cell.value for cell in ws[1]]
    column_by_name = {name: idx + 1 for idx, name in enumerate(header)}

    manual_bool_fields = [
        field
        for field in MANUAL_FIELDS
        if field not in {"人工_最终标签", "人工_错误来源", "人工_备注"}
    ]

    def add_validation(field: str, formula: str) -> None:
        col = column_by_name[field]
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=data_rows + 1, column=col).coordinate}")

    for field in manual_bool_fields:
        add_validation(field, yes_no_formula)
    add_validation("人工_最终标签", label_formula)
    add_validation("人工_错误来源", failure_formula)


def write_xlsx(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P1人工审计"
    ws.append(fieldnames)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    manual_fill = PatternFill("solid", fgColor="FFF2CC")
    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = manual_fill if name in MANUAL_FIELDS else header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        ws.append([row.get(field, "") for field in fieldnames])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    width_by_field = {
        "row_id": 12,
        "sample_id": 12,
        "qa_key": 12,
        "question": 48,
        "baseline_response": 32,
        "gold_answer": 32,
        "evidence": 42,
        "memory_text": 70,
        "suspected_reason": 48,
        "gate_trace_summary": 42,
        "人工_备注": 36,
    }
    for col_idx, field in enumerate(fieldnames, start=1):
        width = width_by_field.get(field, 18 if field in MANUAL_FIELDS else 16)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    add_xlsx_validation(wb, ws, len(rows))
    wb.save(path)


def build_p1_rows(
    candidate_rows: list[dict[str, str]],
    qa_by_idx: dict[int, dict[str, Any]],
    census_by_idx: dict[int, dict[str, Any]],
) -> list[dict[str, Any]]:
    p1 = [row for row in candidate_rows if row.get("priority_group") == P1_GROUP]
    if not p1:
        p1 = [
            row
            for row in candidate_rows
            if row.get("question_type") == "Memory Conflict"
            and row.get("baseline_label") == "omission"
            and not normalize_bool(row.get("r3a_triggered"))
        ]

    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in p1:
        grouped[str(row.get("idx", ""))].append(row)

    ranked: list[dict[str, Any]] = []
    for sample_id, rows in grouped.items():
        rows.sort(
            key=lambda row: (
                -to_float(row.get("total_suspicion_score")),
                to_int(row.get("memory_index")),
            )
        )
        for rank, row in enumerate(rows, start=1):
            sample_idx = to_int(row.get("idx"))
            qa = qa_by_idx.get(sample_idx, {})
            census = census_by_idx.get(sample_idx, {})
            raw_memory_count = census.get("raw_memory_count")
            if raw_memory_count is None and qa.get("raw_memories") is not None:
                raw_memory_count = len(qa.get("raw_memories", []))
            gate_summary = (
                f"r3a_triggered={row.get('r3a_triggered','')}; "
                f"reject_reason={row.get('r3a_rejected_reason_if_available','')}; "
                f"scores=topic:{row.get('topic_overlap_score','')},"
                f"negation:{row.get('negation_or_contrast_score','')},"
                f"temporal:{row.get('temporal_conflict_score','')},"
                f"preference:{row.get('preference_reversal_score','')},"
                f"event:{row.get('event_denial_score','')}"
            )
            out = {
                "sample_id": str(sample_idx),
                "qa_key": row.get("qa_key", qa.get("qa_key", "")),
                "question": row.get("question", qa.get("question", "")),
                "question_type": row.get("question_type", qa.get("question_type", "")),
                "r0_label": row.get("baseline_label", qa.get("baseline_label", "")),
                "baseline_response": qa.get("baseline_response", ""),
                "gold_answer": qa.get("gold_answer", ""),
                "evidence": compact_json(qa.get("evidence", "")),
                "memory_index": row.get("memory_index", ""),
                "memory_text": row.get("memory_text", ""),
                "suspected_reason": row.get("suspected_reason", ""),
                "candidate_score": row.get("total_suspicion_score", ""),
                "r3a_triggered": row.get("r3a_triggered", ""),
                "r3a_reject_reason": row.get("r3a_rejected_reason_if_available", ""),
                "gate_trace_summary": gate_summary,
                "raw_memory_count": raw_memory_count if raw_memory_count is not None else "",
                "candidate_rank_within_sample": rank,
                "same_sample_candidate_count": len(rows),
            }
            for manual_field in MANUAL_FIELDS:
                out[manual_field] = ""
            ranked.append(out)

    ranked.sort(
        key=lambda row: (
            to_int(row.get("sample_id")),
            -to_float(row.get("candidate_score")),
            to_int(row.get("candidate_rank_within_sample")),
        )
    )
    for index, row in enumerate(ranked, start=1):
        row["row_id"] = f"P1-{index:04d}"
    return ranked


def render_label_check_report(check: dict[str, Any]) -> str:
    return f"""# round_0014 标签分布统计口径核查

## 1. 核查对象

round_0014 summary/report 中的 `r3a_untriggered_suspected_baseline_label_distribution`。

## 2. round_0014 已报告分布

```json
{json.dumps(check["reported_distribution"], ensure_ascii=False, indent=2)}
```

该分布总数为 {check["reported_total"]}。

## 3. 重新计算：Memory Conflict + R3a 未触发 + 疑似反证候选

```json
{json.dumps(check["recomputed_memory_conflict_untriggered_suspected_distribution"], ensure_ascii=False, indent=2)}
```

该分布总数为 {check["recomputed_memory_conflict_untriggered_suspected_total"]}。

## 4. 重新计算：全量样本 + R3a 未触发 + 疑似反证候选

```json
{json.dumps(check["recomputed_all_untriggered_suspected_distribution"], ensure_ascii=False, indent=2)}
```

该分布总数为 {check["recomputed_all_untriggered_suspected_total"]}。

## 5. 结论

{check["explanation"]}
"""


def render_prep_report(
    *,
    inputs: dict[str, str],
    outputs: list[str],
    p1_rows: int,
    p1_unique_samples: int,
    check: dict[str, Any],
    row_count_matches: bool,
) -> str:
    mismatch_note = (
        "第一优先级候选行数与 round_0014 预期 120 行一致。"
        if row_count_matches
        else f"第一优先级候选行数为 {p1_rows}，与预期 120 行不一致；需回看 manual_audit_candidates.csv 的抽样逻辑。"
    )
    return f"""# R3a-v2.2 第一优先级人工审计准备报告

## 1. 本轮目的

本轮只做第一优先级疑似反证候选的人工审计表准备，以及 round_0014 标签分布统计口径核查。

未发接口调用，未重新生成回答，未修改 R3a 选择器、提示词、检索器或记忆库。

## 2. 输入文件

- `{inputs["candidates"]}`
- `{inputs["summary"]}`
- `{inputs["report"]}`
- `{inputs["sample_census"]}`
- `{inputs["untriggered_suspected"]}`
- `{inputs["qa"]}`

## 3. 第一优先级筛选规则

从 `manual_audit_candidates.csv` 中筛选：

`priority_group = P1_memory_conflict_omission_untriggered_suspected`

该优先级含义是：Memory Conflict、R0 omission、R3a 未触发、存在疑似反证候选。疑似候选仅用于人工审计，不是真反证标签，也不是应触发标签。

## 4. 第一优先级候选行数和唯一样本数

- 第一优先级候选行数：{p1_rows}
- 第一优先级唯一样本数：{p1_unique_samples}
- 预期行数：120
- 行数是否匹配：{str(row_count_matches).lower()}

{mismatch_note}

## 5. 人工审计表字段说明

审计表包含自动字段和人工填写字段。自动字段来自 round_0014 候选表、全量样本普查、以及原始 QA 文件；人工字段保持空白，供后续人工填写。

`人工_最终标签` 只能使用：真反证、模糊反证、非反证但有用、非反证、无法判断。

`人工_错误来源` 只能使用：R3a漏触发、候选发现器误报、问题前提抽取错误、槽位归类错误、时间判断问题、明确反驳判断问题、检索证据不足、不属于R3a、无法判断。

## 6. 统计口径核查

- round_0014 已报告分布：`{check["reported_distribution"]}`
- Memory Conflict + 未触发 + 疑似候选重新计算：`{check["recomputed_memory_conflict_untriggered_suspected_distribution"]}`
- 全量未触发 + 疑似候选重新计算：`{check["recomputed_all_untriggered_suspected_distribution"]}`
- round_0014 字段名/报告位置是否正确：{str(check["is_round_0014_label_distribution_name_correct"]).lower()}

结论：{check["explanation"]}

## 7. 当前不能得出的结论

不能说第一优先级候选是真反证。
不能说 R3a 漏触发。
不能说 R3a 应该修改。
不能说 R3a 已成功。
不能根据候选发现器直接扩大实验。

## 8. 下一步人工填写说明

请人工填写 `p1_manual_audit_sheet.xlsx`，重点判断候选记忆是否明确反驳问题前提，以及是否应该触发 R3a。

如果 P1 中“真反证且应该触发”的比例很高：说明 R3a 在最重要的遗漏修复场景中漏触发，下一步应修 R3a。

如果 P1 中大多数是“非反证”或“只是相关”：说明高召回候选发现器误报高，R3a 暂不应修改。

如果 P1 中大量是“模糊反证”：说明需要定义边界，不要直接放宽选择器。

如果 P1 中大量属于时间更新：应交给 R3b，不要强行塞给 R3a。

## 9. 生成文件列表

{chr(10).join(f"- `{path}`" for path in outputs)}
"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--candidates", required=True)
    parser.add_argument("--summary", required=True)
    parser.add_argument("--sample-census", required=True)
    parser.add_argument("--qa", required=True)
    parser.add_argument("--outdir", required=True)
    parser.add_argument("--report", default="")
    parser.add_argument("--untriggered-suspected", default="")
    args = parser.parse_args()

    candidates_path = Path(args.candidates)
    summary_path = Path(args.summary)
    sample_census_path = Path(args.sample_census)
    qa_path = Path(args.qa)
    outdir = Path(args.outdir)
    report_path = Path(args.report) if args.report else summary_path.with_name("full_opportunity_census_report.md")
    untriggered_path = (
        Path(args.untriggered_suspected)
        if args.untriggered_suspected
        else summary_path.with_name("r3a_untriggered_suspected_counterevidence.csv")
    )
    outdir.mkdir(parents=True, exist_ok=True)

    candidate_rows = load_csv_rows(candidates_path)
    summary = load_json(summary_path)
    sample_census = load_jsonl_by_idx(sample_census_path)
    qa_by_idx = load_qa_by_idx(qa_path)
    untriggered_rows = load_csv_rows(untriggered_path)
    if report_path.exists():
        _ = report_path.read_text(encoding="utf-8")

    p1_rows = build_p1_rows(candidate_rows, qa_by_idx, sample_census)
    fieldnames = AUTO_FIELDS + MANUAL_FIELDS

    p1_csv = outdir / "p1_manual_audit_sheet.csv"
    p1_xlsx = outdir / "p1_manual_audit_sheet.xlsx"
    write_csv(p1_csv, p1_rows, fieldnames)
    write_xlsx(p1_xlsx, p1_rows, fieldnames)

    reported = summary.get("r3a_untriggered_suspected_baseline_label_distribution", {})
    reported_total = sum(int(v) for v in reported.values())
    mc_dist = plain_counter(
        distribution_from_sample_census(sample_census, memory_conflict_only=True)
    )
    all_dist = plain_counter(
        distribution_from_sample_census(sample_census, memory_conflict_only=False)
    )
    csv_all_dist = plain_counter(distribution_from_untriggered_csv(untriggered_rows))
    reported_matches_mc = reported == mc_dist
    reported_matches_all = reported == all_dist
    csv_matches_all = csv_all_dist == all_dist

    if reported_matches_all and not reported_matches_mc:
        label_issue = (
            "round_0014 中该字段实际对应全量未触发且存在疑似候选的 baseline_label 分布，"
            "不是 Memory Conflict 子集分布；字段名或报告中的上下文应修正为“全量未触发疑似候选 baseline_label 分布”。"
        )
        name_correct = False
    elif reported_matches_mc:
        label_issue = "round_0014 中该字段与 Memory Conflict 子集口径一致。"
        name_correct = True
    else:
        label_issue = (
            "round_0014 中该字段既不匹配 Memory Conflict 子集，也不匹配全量未触发疑似候选口径；"
            "后续需要修正普查脚本统计逻辑。"
        )
        name_correct = False

    label_check = {
        "reported_distribution": reported,
        "reported_total": reported_total,
        "recomputed_memory_conflict_untriggered_suspected_distribution": mc_dist,
        "recomputed_memory_conflict_untriggered_suspected_total": sum(mc_dist.values()),
        "recomputed_all_untriggered_suspected_distribution": all_dist,
        "recomputed_all_untriggered_suspected_total": sum(all_dist.values()),
        "csv_all_untriggered_suspected_distribution": csv_all_dist,
        "csv_all_untriggered_suspected_total": sum(csv_all_dist.values()),
        "reported_matches_memory_conflict_untriggered_suspected": reported_matches_mc,
        "reported_matches_all_untriggered_suspected": reported_matches_all,
        "csv_matches_sample_census_all_untriggered_suspected": csv_matches_all,
        "is_round_0014_label_distribution_name_correct": name_correct,
        "explanation": label_issue,
    }

    row_count_matches = len(p1_rows) == EXPECTED_P1_ROWS
    unique_samples = len({row["sample_id"] for row in p1_rows})

    summary_out = {
        "round": ROUND,
        "input_round": INPUT_ROUND,
        "task": "prepare_p1_manual_audit_and_check_label_distribution",
        "p1_candidate_rows": len(p1_rows),
        "p1_unique_samples": unique_samples,
        "p1_expected_rows": EXPECTED_P1_ROWS,
        "p1_row_count_matches_round_0014": row_count_matches,
        "api_calls": 0,
        "changed_selector_logic": False,
        "changed_prompt": False,
        "changed_retriever": False,
        "changed_memory_store": False,
        "label_distribution_check": label_check,
        "next_step": "human_fill_p1_manual_audit_sheet",
    }

    label_json = outdir / "round_0014_label_distribution_check.json"
    label_md = outdir / "round_0014_label_distribution_check.md"
    prep_json = outdir / "p1_audit_prep_summary.json"
    prep_md = outdir / "p1_audit_prep_report.md"

    prep_json.write_text(json.dumps(summary_out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    label_json.write_text(json.dumps(label_check, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    label_md.write_text(render_label_check_report(label_check), encoding="utf-8")

    outputs = [
        str(p1_xlsx),
        str(p1_csv),
        str(prep_json),
        str(prep_md),
        str(label_md),
        str(label_json),
    ]
    report_text = render_prep_report(
        inputs={
            "candidates": str(candidates_path),
            "summary": str(summary_path),
            "report": str(report_path),
            "sample_census": str(sample_census_path),
            "untriggered_suspected": str(untriggered_path),
            "qa": str(qa_path),
        },
        outputs=outputs,
        p1_rows=len(p1_rows),
        p1_unique_samples=unique_samples,
        check=label_check,
        row_count_matches=row_count_matches,
    )
    prep_md.write_text(report_text, encoding="utf-8")

    print(json.dumps(summary_out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
